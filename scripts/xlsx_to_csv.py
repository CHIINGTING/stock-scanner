#!/usr/bin/env python3
"""xlsx_to_csv.py — 把官方 XLSX 的每個工作表轉成 CSV。

用法:
    python3 scripts/xlsx_to_csv.py data/official/<file.xlsx>

輸出:
    data/official/csv/<原檔名>__<工作表名>.csv

特性:
    - 使用 openpyxl 讀取（data_only=True，取公式計算結果）。
    - 每個 worksheet 各輸出一個 CSV。
    - UTF-8 with BOM（utf-8-sig），避免 Excel 開啟中文亂碼。
    - 跳過完全空白的列。
"""

import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write("需要 openpyxl：python3 -m pip install openpyxl\n")
    sys.exit(1)

OUT_DIR = Path("data/official/csv")


def safe_name(s: str) -> str:
    """把工作表名清成安全的檔名片段。"""
    keep = []
    for ch in s.strip():
        keep.append(ch if (ch.isalnum() or ch in "-_") else "_")
    return "".join(keep) or "sheet"


def is_blank_row(row) -> bool:
    return all(c is None or str(c).strip() == "" for c in row)


def convert(xlsx_path: Path) -> list[Path]:
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"找不到輸入檔：{xlsx_path}")

    wb = load_workbook(filename=xlsx_path, data_only=True, read_only=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    stem = xlsx_path.stem
    written: list[Path] = []

    for ws in wb.worksheets:
        out_path = OUT_DIR / f"{stem}__{safe_name(ws.title)}.csv"
        rows_written = 0
        # utf-8-sig => UTF-8 with BOM
        with out_path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.writer(fh)
            for row in ws.iter_rows(values_only=True):
                if is_blank_row(row):
                    continue
                writer.writerow(["" if v is None else v for v in row])
                rows_written += 1
        if rows_written == 0:
            out_path.unlink(missing_ok=True)  # 不留空白 CSV
            print(f"  跳過空白工作表：{ws.title}")
            continue
        written.append(out_path)
        print(f"  {ws.title} -> {out_path} ({rows_written} 列)")

    wb.close()
    return written


def main() -> int:
    if len(sys.argv) != 2:
        sys.stderr.write("用法: python3 scripts/xlsx_to_csv.py <file.xlsx>\n")
        return 2
    xlsx_path = Path(sys.argv[1])
    try:
        written = convert(xlsx_path)
    except Exception as e:  # noqa: BLE001 — CLI 邊界，回報清楚訊息即可
        sys.stderr.write(f"轉換失敗：{e}\n")
        return 1
    if not written:
        sys.stderr.write("沒有任何非空白工作表可輸出。\n")
        return 1
    print(f"完成，輸出 {len(written)} 個 CSV 到 {OUT_DIR}/")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
